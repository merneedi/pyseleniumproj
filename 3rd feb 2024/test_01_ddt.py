# username, password
# admin, admin
# admin , admin123
# ...
# ..
# etc

import pytest
from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
import time

# Read the Excel the get the data
def get_test_data():
    workbook = load_workbook("test_data.xlsx")
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(min_row =2, values_only = True):
        data.append(row)
    return data

@pytest.fixture
def driver():
    driver = webdriver.Chrome()
    driver.get('https://app.vwo.com')
    driver.maximize_window()
    yield driver
    driver.quit()

@pytest.mark.parametrize("username,password,result",get_test_data())
def test_vwo_login(driver,username,password,result):
    email =driver.find_element(By.ID,"login-username")
    email.send_keys(username)

    pass_wrd = driver.find_element(By.ID, "login-password")
    pass_wrd.send_keys(password)

    driver.find_element(By.ID, "js-login-btn").click()
    time.sleep(3)
    # username, password - 1 cort username, password
    # navigate to url
    # enter the username and password
    # click
    # verify if the login is working yes or no
    print(username,password,driver.current_url)

    if result == "Fail":
        error_msg = driver.find_element(By.ID,"js-notification-box-msg").text
        assert error_msg in "Your email, password, IP address or location did not match"

    else:
        assert "dashboard" in driver.current_url